# coding=utf-8

# Builds a compact inventory report for the spools tab in three formats: PDF, CSV, XLSX.
# Columns: Material, Color, Vendor, Remaining weight and - only if any spool has notes - Notes.
# All three formatters share the same column/row extraction (extract_report_rows) so the
# report content stays identical across formats. See GitHub issue #209 / #141.

import csv
import datetime
import socket
from io import BytesIO, StringIO

from octoprint_SpoolManager.api.Transformer import calculateRemainingWeight

FORMAT_DATETIME = "%d.%m.%Y %H:%M"

# Fixed columns; "Notes" is appended dynamically only when at least one spool has notes.
BASE_COLUMNS = ["Material", "Color", "Vendor", "Remaining"]
NOTES_COLUMN = "Notes"

LOOPBACK_HOSTS = ("localhost", "127.0.0.1", "::1")


def _value(rawValue):
    if (rawValue is None):
        return ""
    return str(rawValue)


def _resolveHost(host):
    """
    Resolves a loopback host ("localhost"/"127.0.0.1") to this machine's actual LAN IP,
    so the report stays meaningful when the external database happens to run on the same
    host as this OctoPrint instance. Falls back to the original value if resolution fails.
    """
    if (host not in LOOPBACK_HOSTS):
        return host

    try:
        probeSocket = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            # Doesn't actually send anything (UDP) - just makes the OS pick the outbound
            # interface/IP that would be used to reach an external address.
            probeSocket.connect(("8.8.8.8", 80))
            return probeSocket.getsockname()[0]
        finally:
            probeSocket.close()
    except OSError:
        return host


def _formatRemainingWeight(spoolModel):
    remaining = calculateRemainingWeight(spoolModel.usedWeight, spoolModel.totalWeight)
    if (remaining is None):
        return "-"
    return "%.0f g" % remaining


def build_database_context_info(databaseSettings, instanceName):
    """
    Builds a short human-readable string identifying which database the report was
    generated from - external DB: host + database name; internal DB: OctoPrint
    instance name (Appearance -> Title) + the fixed internal database file name.

    :param databaseSettings: DatabaseManager.DatabaseSettings (or None)
    :param instanceName: OctoPrint instance name from Appearance settings (may be None/empty)
    :return: str
    """
    if (databaseSettings is not None and databaseSettings.useExternal == True):
        host = _resolveHost(_value(databaseSettings.host))
        return "%s / %s" % (host, _value(databaseSettings.name))

    name = instanceName if (instanceName is not None and str(instanceName).strip() != "") else "OctoPrint"
    return "%s / spoolmanager.db" % name


def extract_report_rows(allSpoolModels):
    """
    Extracts the report header and rows from the given (already filtered/sorted) spool models.
    The Notes column is only included when at least one spool actually has notes.

    :param allSpoolModels: iterable of SpoolModel (may be None/empty)
    :return: tuple (headers: list[str], rows: list[list[str]], hasNotes: bool)
    """
    spools = list(allSpoolModels) if allSpoolModels is not None else []

    hasNotes = any((_value(spool.noteText).strip() != "") for spool in spools)

    headers = list(BASE_COLUMNS)
    if (hasNotes):
        headers.append(NOTES_COLUMN)

    rows = []
    for spool in spools:
        row = [
            _value(spool.material),
            _value(spool.colorName),
            _value(spool.vendor),
            _formatRemainingWeight(spool),
        ]
        if (hasNotes):
            row.append(_value(spool.noteText))
        rows.append(row)

    return headers, rows, hasNotes


####################################################################################### -> PDF

def build_inventory_report_pdf(allSpoolModels, dbContextInfo=""):
    """Builds the inventory report as a landscape A4 PDF. Returns a BytesIO positioned at 0."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    headers, rows, hasNotes = extract_report_rows(allSpoolModels)

    styles = getSampleStyleSheet()
    cellStyle = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=10)
    headerStyle = ParagraphStyle("cellHeader", parent=cellStyle, textColor=colors.white)

    tableData = [[Paragraph(h, headerStyle) for h in headers]]
    for row in rows:
        tableData.append([Paragraph(cell, cellStyle) for cell in row])

    # Column widths tuned for A4 landscape (usable width ~277mm after margins).
    if (hasNotes):
        colWidths = [40 * mm, 40 * mm, 45 * mm, 30 * mm, 122 * mm]
    else:
        colWidths = [55 * mm, 55 * mm, 90 * mm, 77 * mm]

    table = Table(tableData, colWidths=colWidths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bdc3c7")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f7")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    now = datetime.datetime.now()
    titleStyle = ParagraphStyle("title", parent=styles["Title"], fontSize=16, spaceAfter=2)
    subTitleStyle = ParagraphStyle("subTitle", parent=styles["Normal"], fontSize=9,
                                   textColor=colors.HexColor("#7f8c8d"))
    dbInfoStyle = ParagraphStyle("dbInfo", parent=styles["Normal"], fontSize=8,
                                 textColor=colors.HexColor("#7f8c8d"), alignment=TA_RIGHT)

    titleCell = [
        Paragraph("SpoolManager Inventory Report", titleStyle),
        Paragraph("%d spool(s) &middot; generated %s" % (len(rows), now.strftime(FORMAT_DATETIME)),
                  subTitleStyle),
    ]
    dbInfoCell = Paragraph(dbContextInfo, dbInfoStyle)

    # 1-row, borderless table as a layout trick to place the DB info top-right next to the title.
    headerTable = Table([[titleCell, dbInfoCell]], colWidths=[187 * mm, 90 * mm])
    headerTable.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    story = [
        headerTable,
        Spacer(1, 6 * mm),
        table,
    ]

    pdfBuffer = BytesIO()
    doc = SimpleDocTemplate(
        pdfBuffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="SpoolManager Inventory Report",
    )
    doc.build(story)
    pdfBuffer.seek(0)
    return pdfBuffer


####################################################################################### -> CSV

def build_inventory_report_csv(allSpoolModels, dbContextInfo=""):
    """Builds the inventory report as CSV (comma-delimited, all fields quoted). Returns str."""
    headers, rows, _ = extract_report_rows(allSpoolModels)

    output = StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL, lineterminator="\n")
    # Comment line (# prefix) + blank separator, so the data grid below stays a clean
    # rectangular table for spreadsheet import (Excel/LibreOffice column auto-detection).
    writer.writerow(["# SpoolManager Inventory Report - " + dbContextInfo])
    writer.writerow([])
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


####################################################################################### -> XLSX

def build_inventory_report_xlsx(allSpoolModels, dbContextInfo=""):
    """Builds the inventory report as an .xlsx workbook. Returns a BytesIO positioned at 0."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    headers, rows, _ = extract_report_rows(allSpoolModels)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    titleFont = Font(bold=True)
    ws.append(["SpoolManager Inventory Report", dbContextInfo])
    ws["A1"].font = titleFont
    ws.append([])

    headerFont = Font(bold=True, color="FFFFFF")
    headerFill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")

    headerRowIndex = 3
    ws.append(headers)
    for cell in ws[headerRowIndex]:
        cell.font = headerFont
        cell.fill = headerFill

    for row in rows:
        ws.append(row)

    # Auto-size columns to the widest cell (header included), with a sane cap.
    for columnIndex, header in enumerate(headers, start=1):
        maxLength = len(header)
        for row in rows:
            value = row[columnIndex - 1]
            if (len(value) > maxLength):
                maxLength = len(value)
        ws.column_dimensions[ws.cell(row=headerRowIndex, column=columnIndex).column_letter].width = min(maxLength + 2, 60)

    xlsxBuffer = BytesIO()
    wb.save(xlsxBuffer)
    xlsxBuffer.seek(0)
    return xlsxBuffer
